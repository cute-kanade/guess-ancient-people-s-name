#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 peoples_names.md 转换为 peoples_names.xls（实际为 xlsx 格式）
首列为序号，第二列为姓名
"""

import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def convert_md_to_xls(input_file: str, output_file: str) -> None:
    # 读取 Markdown 文件
    with open(input_file, encoding="utf-8") as f:
        lines = f.readlines()

    # 解析每一行，格式为 "序号. 姓名"
    records = []
    pattern = re.compile(r"^(\d+)\.\s+(.+)$")
    for line in lines:
        line = line.strip()
        if not line:
            continue
        m = pattern.match(line)
        if m:
            index = int(m.group(1))
            name  = m.group(2).strip()
            records.append((index, name))
        else:
            print(f"  [跳过] 无法解析的行: {line!r}")

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "人物名单"

    # 样式定义
    header_font  = Font(bold=True, size=12, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2F5496")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写表头
    for col, header in enumerate(["序号", "姓名"], start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = cell_border

    # 冻结首行
    ws.freeze_panes = "A2"

    # 写数据
    for row_idx, (index, name) in enumerate(records, start=2):
        # 序号列
        c0 = ws.cell(row=row_idx, column=1, value=index)
        c0.alignment = center_align
        c0.border    = cell_border

        # 姓名列
        c1 = ws.cell(row=row_idx, column=2, value=name)
        c1.alignment = left_align
        c1.border    = cell_border

        # 隔行底色
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="EEF3FB")
            c0.fill = fill
            c1.fill = fill

    # 设置列宽
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22

    # 行高
    ws.row_dimensions[1].height = 22
    for r in range(2, len(records) + 2):
        ws.row_dimensions[r].height = 18

    wb.save(output_file)
    print(f"✅ 转换完成！共写入 {len(records)} 条记录。")
    print(f"   输出文件: {os.path.abspath(output_file)}")


if __name__ == "__main__":
    base_dir    = os.path.dirname(os.path.abspath(__file__))
    input_file  = os.path.join(base_dir, "peoples_names.md")
    output_file = os.path.join(base_dir, "peoples_names.xls")
    convert_md_to_xls(input_file, output_file)
